#!/usr/bin/env python3
"""
Download ALL Excel support files and parse them for vehicle/ECU coverage data.
Phase 1: Extract Excel URLs from HTML, download missing ones.
Phase 2: Parse all downloaded Excel files for structured data.
"""

import os
import re
import json
import time
import glob
import urllib.request
import urllib.error
from pathlib import Path
from bs4 import BeautifulSoup
from collections import defaultdict

HTML_DIR = Path("/Users/jeremysamuels/Documents/study-dashboard/data/obdii365_scraped/html")
DOWNLOAD_DIR = Path("/Users/jeremysamuels/Documents/study-dashboard/data/obdii365_scraped/support_files")
OUTPUT_DIR = Path("/Users/jeremysamuels/Documents/study-dashboard/data/obdii365_scraped/parsed")
BASE_URL = "https://www.obdii365.com"


def extract_excel_links():
    """Extract all Excel file links from HTML pages."""
    html_files = sorted(HTML_DIR.glob("*.html"))
    html_files = [f for f in html_files if not f.name.startswith("listing_")]
    
    all_links = []
    for filepath in html_files:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        soup = BeautifulSoup(content, 'html.parser')
        h1 = soup.find('h1')
        product_name = h1.get_text(strip=True) if h1 else filepath.stem
        
        for link in soup.find_all('a', href=True):
            href = link.get('href', '').strip()
            text = link.get_text(strip=True)
            
            if not href:
                continue
            
            href_lower = href.lower()
            if '.xlsx' not in href_lower and '.xls' not in href_lower:
                continue
            if '.xls' in href_lower and '.xlsx' not in href_lower:
                ext = '.xls'
            else:
                ext = '.xlsx'
            
            if href.startswith('/'):
                full_url = BASE_URL + href
            elif href.startswith('http'):
                full_url = href
            else:
                full_url = BASE_URL + '/' + href
            
            # Create descriptive filename
            if '/' in href:
                raw_name = href.rstrip('/').split('/')[-1]
            else:
                raw_name = href
            
            # Make sure extension is present   
            if not any(raw_name.lower().endswith(e) for e in ['.xlsx', '.xls']):
                raw_name += ext
            
            # Prefix with product slug for context
            product_slug = re.sub(r'[^\w]', '_', filepath.stem)[:40]
            
            all_links.append({
                'product_file': filepath.name,
                'product_name': product_name[:120],
                'product_slug': product_slug,
                'link_text': text[:200],
                'url': full_url,
                'raw_filename': raw_name,
            })
    
    # Deduplicate by URL
    seen = set()
    unique = []
    for link in all_links:
        if link['url'] not in seen:
            seen.add(link['url'])
            unique.append(link)
    
    return unique


def download_file(url, dest_path, retries=2):
    """Download a file."""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': '*/*',
        'Referer': 'https://www.obdii365.com/',
    }
    
    for attempt in range(retries + 1):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=30) as resp:
                content = resp.read()
                if len(content) < 100:
                    return False, "File too small"
                with open(dest_path, 'wb') as f:
                    f.write(content)
                return True, len(content)
        except urllib.error.HTTPError as e:
            if e.code == 404:
                return False, "404"
            if attempt < retries:
                time.sleep(1)
        except Exception as e:
            if attempt < retries:
                time.sleep(1)
            else:
                return False, str(e)[:80]
    return False, "retries exceeded"


def parse_excel_file(filepath):
    """Parse a single Excel file for vehicle/ECU coverage data."""
    try:
        import openpyxl
    except ImportError:
        # Fallback: try pandas
        try:
            import pandas as pd
            return parse_excel_with_pandas(filepath)
        except ImportError:
            return None
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return {"error": str(e)[:100], "filepath": str(filepath)}
    
    result = {
        "filepath": str(filepath),
        "filename": filepath.name,
        "sheets": [],
        "total_rows": 0,
        "vehicle_entries": [],
        "ecu_entries": [],
        "function_entries": [],
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        sheet_data = {
            "name": sheet_name,
            "rows": 0,
            "columns": [],
            "sample_rows": [],
        }
        
        rows = list(ws.iter_rows(values_only=True))
        sheet_data["rows"] = len(rows)
        result["total_rows"] += len(rows)
        
        if not rows:
            result["sheets"].append(sheet_data)
            continue
        
        # Get headers from first row
        headers = []
        for cell in rows[0]:
            if cell is not None:
                headers.append(str(cell).strip())
            else:
                headers.append("")
        sheet_data["columns"] = headers
        
        # Sample first 5 data rows
        for row in rows[1:6]:
            row_data = {}
            for j, cell in enumerate(row):
                if j < len(headers) and headers[j] and cell is not None:
                    row_data[headers[j]] = str(cell).strip()
            if row_data:
                sheet_data["sample_rows"].append(row_data)
        
        # Parse ALL rows for vehicle/ECU data
        header_lower = [h.lower() for h in headers]
        
        # Detect column types
        make_col = next((i for i, h in enumerate(header_lower) if any(kw in h for kw in ['make', 'brand', 'manufacturer', 'car'])), None)
        model_col = next((i for i, h in enumerate(header_lower) if any(kw in h for kw in ['model', 'vehicle'])), None)
        year_col = next((i for i, h in enumerate(header_lower) if any(kw in h for kw in ['year', 'from', 'start'])), None)
        year_end_col = next((i for i, h in enumerate(header_lower) if any(kw in h for kw in ['to', 'end'])), None)
        func_col = next((i for i, h in enumerate(header_lower) if any(kw in h for kw in ['function', 'support', 'feature', 'type', 'key type', 'key_type'])), None)
        ecu_col = next((i for i, h in enumerate(header_lower) if any(kw in h for kw in ['ecu', 'module', 'system', 'chip', 'transponder'])), None)
        
        for row in rows[1:]:
            entry = {}
            
            if make_col is not None and make_col < len(row) and row[make_col]:
                entry['make'] = str(row[make_col]).strip()
            if model_col is not None and model_col < len(row) and row[model_col]:
                entry['model'] = str(row[model_col]).strip()
            if year_col is not None and year_col < len(row) and row[year_col]:
                entry['year_start'] = str(row[year_col]).strip()
            if year_end_col is not None and year_end_col < len(row) and row[year_end_col]:
                entry['year_end'] = str(row[year_end_col]).strip()
            if func_col is not None and func_col < len(row) and row[func_col]:
                entry['function'] = str(row[func_col]).strip()
            if ecu_col is not None and ecu_col < len(row) and row[ecu_col]:
                entry['ecu'] = str(row[ecu_col]).strip()
            
            if entry.get('make') or entry.get('model'):
                result["vehicle_entries"].append(entry)
            elif entry.get('ecu'):
                result["ecu_entries"].append(entry)
            elif entry.get('function'):
                result["function_entries"].append(entry)
        
        result["sheets"].append(sheet_data)
    
    wb.close()
    return result


def parse_xls_file(filepath):
    """Parse .xls (old format) files using xlrd."""
    try:
        import xlrd
    except ImportError:
        return {"error": "xlrd not installed", "filepath": str(filepath)}
    
    try:
        wb = xlrd.open_workbook(filepath)
    except Exception as e:
        return {"error": str(e)[:100], "filepath": str(filepath)}
    
    result = {
        "filepath": str(filepath),
        "filename": filepath.name,
        "sheets": [],
        "total_rows": 0,
        "vehicle_entries": [],
        "ecu_entries": [],
        "function_entries": [],
    }
    
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        
        sheet_data = {
            "name": sheet_name,
            "rows": ws.nrows,
            "columns": [],
            "sample_rows": [],
        }
        result["total_rows"] += ws.nrows
        
        if ws.nrows == 0:
            result["sheets"].append(sheet_data)
            continue
        
        # Headers
        headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
        sheet_data["columns"] = headers
        
        # Sample rows
        for r in range(1, min(6, ws.nrows)):
            row_data = {}
            for c in range(ws.ncols):
                if headers[c] and ws.cell_value(r, c):
                    row_data[headers[c]] = str(ws.cell_value(r, c)).strip()
            if row_data:
                sheet_data["sample_rows"].append(row_data)
        
        result["sheets"].append(sheet_data)
    
    return result


def main():
    print("=" * 60)
    print("OBDII365 Excel Support File Pipeline")
    print("=" * 60)
    
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    excel_dir = DOWNLOAD_DIR / "excel"
    excel_dir.mkdir(parents=True, exist_ok=True)
    
    # Phase 1: Extract and download Excel files
    print("\nPhase 1: Extracting Excel URLs from HTML pages...")
    links = extract_excel_links()
    print(f"Found {len(links)} unique Excel file URLs")
    
    # Show what products they belong to
    products_with_excel = set(l['product_name'] for l in links)
    print(f"From {len(products_with_excel)} unique products")
    
    downloaded = 0
    skipped = 0
    failed = 0
    download_manifest = []
    
    for i, link in enumerate(links, 1):
        # Use product slug + raw filename for uniqueness
        safe_name = re.sub(r'[^\w.-]', '_', link['raw_filename'])[:80]
        # Prefix with a unique identifier from URL to avoid collisions
        url_id = re.search(r'(\d{10,})', link['url'])
        if url_id:
            safe_name = url_id.group(1)[:15] + '_' + safe_name
        
        dest_path = excel_dir / safe_name
        
        if dest_path.exists() and dest_path.stat().st_size > 100:
            skipped += 1
            link['local_path'] = str(dest_path)
            download_manifest.append(link)
            continue
        
        # Also check in relevant/ dir from earlier download
        alt_path = DOWNLOAD_DIR / 'relevant' / link['raw_filename']
        if alt_path.exists() and alt_path.stat().st_size > 100:
            # Copy to excel dir
            import shutil
            shutil.copy2(alt_path, dest_path)
            skipped += 1
            link['local_path'] = str(dest_path)
            download_manifest.append(link)
            continue
        
        print(f"  [{i}/{len(links)}] Downloading: {safe_name[:60]} ({link['product_slug'][:30]})")
        success, result = download_file(link['url'], dest_path)
        
        if success:
            downloaded += 1
            link['local_path'] = str(dest_path)
            link['size'] = result
            download_manifest.append(link)
        else:
            failed += 1
            if dest_path.exists():
                dest_path.unlink()
        
        time.sleep(0.3)
    
    print(f"\nDownloads: {downloaded} new, {skipped} existing, {failed} failed")
    
    # Phase 2: Parse all downloaded Excel files
    print(f"\nPhase 2: Parsing Excel files...")
    
    excel_files = list(excel_dir.glob("*.xlsx")) + list(excel_dir.glob("*.xls"))
    # Also include files from relevant/ dir
    excel_files += list((DOWNLOAD_DIR / 'relevant').glob("*.xlsx"))
    excel_files += list((DOWNLOAD_DIR / 'relevant').glob("*.xls"))
    
    # Deduplicate
    seen = set()
    unique_files = []
    for f in excel_files:
        if f.name not in seen and f.stat().st_size > 100:
            seen.add(f.name)
            unique_files.append(f)
    
    print(f"Found {len(unique_files)} unique Excel files to parse")
    
    all_parsed = []
    total_vehicle_entries = 0
    total_ecu_entries = 0
    
    for filepath in sorted(unique_files):
        print(f"  Parsing: {filepath.name}")
        
        if filepath.suffix == '.xls':
            parsed = parse_xls_file(filepath)
        else:
            parsed = parse_excel_file(filepath)
        
        if parsed:
            # Look up which product this came from
            matching = [l for l in download_manifest if 
                       filepath.name in l.get('local_path', '') or 
                       filepath.name == l['raw_filename']]
            if matching:
                parsed['product_name'] = matching[0]['product_name']
                parsed['product_file'] = matching[0]['product_file']
                parsed['link_text'] = matching[0]['link_text']
            
            all_parsed.append(parsed)
            ve = len(parsed.get('vehicle_entries', []))
            ee = len(parsed.get('ecu_entries', []))
            total_vehicle_entries += ve
            total_ecu_entries += ee
            
            if ve > 0 or ee > 0:
                print(f"    -> {ve} vehicle entries, {ee} ECU entries")
            
            for sheet in parsed.get('sheets', []):
                if sheet['columns']:
                    print(f"    Sheet '{sheet['name']}': {sheet['rows']} rows, cols: {sheet['columns'][:5]}")
    
    # Save parsed results
    output = {
        "total_files": len(unique_files),
        "total_vehicle_entries": total_vehicle_entries,
        "total_ecu_entries": total_ecu_entries,
        "files": all_parsed,
    }
    
    output_path = OUTPUT_DIR / "excel_support_data.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, default=str)
    
    print(f"\n{'=' * 60}")
    print(f"Pipeline Complete!")
    print(f"{'=' * 60}")
    print(f"Excel files parsed: {len(unique_files)}")
    print(f"Vehicle entries extracted: {total_vehicle_entries}")
    print(f"ECU entries extracted: {total_ecu_entries}")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
